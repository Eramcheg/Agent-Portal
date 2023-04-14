import urllib.request
import time
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.views import View
from openpyxl import load_workbook
from users.forms import UserCreationForm


class All_tables(View):
    def get(self, request):
        template_name = str(request.user.username)+'_Global_'
        language_code = request.path.split('/')[1]
        context = {'is_home_page': False,
                   'is_tables_page': False,
                   'is_profile_page': False,
                   'is_all_tables_page': True,
                   'is_authenticated': request.user.username == 'admin',
                   'language' : language_code,
                   }

        try:
            return render(request,'admin_Global'+'.html', context)
        except:
            trueOrFalse = self.create_table(template_name)

            if trueOrFalse:
                return render(request, template_name + '.html')
            else:
                return render(request, 'NotFoundHtml.html')

    def create_table(self, name_html):
        try:
            link = 'ftp://oliverweber%40agentsoliverweber.com:Zh5%5DMVF%28GhZ%7B@server1.agentsoliverweber.com:21//ftp/' + f'{name_html}.xlsx'

            with urllib.request.urlopen(link) as response, open(f"{name_html}.xlsx", 'wb') as out_file:
                data = response.read()  # Read the contents of the file
                out_file.write(data)  # Write the contents to the local file

            wb = load_workbook(f'{name_html}.xlsx')
            ws = wb['Sheet']
            html = """<meta charset="utf-8">
                <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" integrity="sha384-Gn5384xqQ1aoWXA+058RXPxPg6fy4IWvTNh0E263XmFcJlSAwiGgFAW/dAiS6JXm" crossorigin="anonymous">
            <!--<link rel="stylesheet" type="text/css" href="normas.css">-->
            <style>
                body{
                    background-color: #d6e1f8;

                }
                th, td {
                    border: 1px solid #ccc;

                    text-align: center;
                }

                table {

                    border-collapse: collapse !important; /* optional: collapse the table borders */
                }

                table tr {
                    background-color: #8BC34A !important; /* Set background color of all rows */
                }

                table tr:hover td {
                    background-color: #2196F3 !important; /* Set background color of cells when hovered */
                }

            </style>
                             <table border="1" class="dataframe" style="width:100%;">
                                         <thead>
                                             <tr style="text-align: right;">\n"""
            k = 0
            counter_max = 0
            for i in ws.iter_rows():
                counter_max += 1

            for i in ws.iter_rows():
                if str(i[0].fill.start_color.rgb) != 'FFFFFFFF' or k == 0 or k == counter_max:
                    if k > 0:
                        html += '<tr>\n'
                    for j in range(len(i) - 1):
                        value = str(i[j].value)
                        color = str(i[j].fill.start_color.rgb)
                        if (value == 'None'):
                            value = ' '
                        if color == '00000000':
                            color = 'FFFFFFFF'

                        if k < 1:
                            html += "<th style='background-color: #" + color[2:8] + "'>" + value + "</th>\n"
                        else:
                            html += "<td style='background-color: #" + color[2:8] + "'>" + value + "</td>\n"
                    html += '</tr>\n'
                    if k < 1:
                        html += """</thead>
                    <tbody>"""
                    k += 1
            html += """ </tbody>
            </table>"""
            html.encode('UTF-8')
            with open('users\\templates\\'+name_html+".html", 'w', encoding='utf-8') as f:
                f.write(html)
                return True
        except:
            return False

class Table(View):
    def get(self, request):
        language_code = request.path.split('/')[2]
        if request.user.is_authenticated:

            username = str(request.user.username)

            print(username)
            context = {'language' : language_code}
            name_html = username + "_" + str(request.GET.get('name'))

            print(name_html)
            html = name_html+'.html'

            try:
                return render(request, html,  context)
            except:
                trueOrFalse = self.create_table(name_html)

                if trueOrFalse:

                    # while not os.path.exists('users\\templates\\'+name_html+'.html'):  # replace "file_path" with the actual path to your file
                    #     time.sleep(1)
                    #     print("File is not ready")
                    # while os.path.isfile("users\\templates\\"+name_html+".html")==False:
                    #     pass
                    return render(request, name_html + '.html', context)
                    # for i in range(10):
                    #     try:
                    #         return render(request, name_html+'.html')
                    #         break
                    #     except:
                    #         time.sleep(6)

                    # for i in range(5):
                        #
                        # try:
                        #     return render(request, name_html+'.html')
                        #     break
                        # except:
                        #     time.sleep(i)


                else:
                      return render(request, 'NotFoundHtml.html',context)
        else:
            return render(request, 'NotFoundHtml.html')
    def create_table(self, name_html):
        try:
            link = 'ftp://oliverweber%40agentsoliverweber.com:Zh5%5DMVF%28GhZ%7B@server1.agentsoliverweber.com:21//ftp/' + f'{name_html}.xlsx'

            with urllib.request.urlopen(link) as response, open(f"{name_html}.xlsx", 'wb') as out_file:
                data = response.read()  # Read the contents of the file
                out_file.write(data)  # Write the contents to the local file

            wb = load_workbook(f'{name_html}.xlsx')
            ws = wb['Sheet']
            html = """{%extends 'base.html'%}
                        {% block content %}
                        <meta charset="utf-8">
                            <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" integrity="sha384-Gn5384xqQ1aoWXA+058RXPxPg6fy4IWvTNh0E263XmFcJlSAwiGgFAW/dAiS6JXm" crossorigin="anonymous">
                        <!--<link rel="stylesheet" type="text/css" href="normas.css">-->
                        <style>
                            body{
                                background-color: #d6e1f8;

                            }
                            th, td {
                                border: 1px solid #ccc;

                                text-align: center;
                            }

                            table {

                                border-collapse: collapse !important; /* optional: collapse the table borders */
                            }

                            table tr {
                                background-color: #8BC34A !important; /* Set background color of all rows */
                            }

                            table tr:hover td {
                                background-color: #2196F3 !important; /* Set background color of cells when hovered */
                            }

                        </style>
                                         <table border="1" class="dataframe" style="width:100%;">
                                                     <thead>
                                                         <tr style="text-align: right;">\n"""
            k = 0
            counter_max = 0
            for i in ws.iter_rows():
                counter_max += 1

            for i in ws.iter_rows():
                if str(i[0].fill.start_color.rgb) != 'FFFFFFFF' or k == 0 or k == counter_max:
                    if k > 0:
                        html += '<tr>\n'
                    for j in range(len(i) - 1):
                        value = str(i[j].value)
                        color = str(i[j].fill.start_color.rgb)
                        if (value == 'None'):
                            value = ' '
                        if color == '00000000':
                            color = 'FFFFFFFF'

                        if k < 1:
                            html += "<th style='background-color: #" + color[2:8] + "'>" + value + "</th>\n"
                        else:
                            html += "<td style='background-color: #" + color[2:8] + "'>" + value + "</td>\n"
                    html += '</tr>\n'
                    if k < 1:
                        html += """</thead>
                            <tbody>"""
                    k += 1

            html += """ </tbody>
                        </table>
                        {% endblock %}
                        """
            html.encode('UTF-8')
            file = open('users\\templates\\'+name_html+'.html', 'w', encoding='utf-8')
            file.write(html)
            file.close()
            # with open(name_html + ".html", 'w', encoding='utf-8') as f:
            #     f.write(html)
            time.sleep(5)
            return True
        except:
            return False

class Home(View):
    def get(self, request):
        language_code = request.path.split('/')[1]
        print(language_code)
        is_auth = request.user.username == 'admin'
        context = {'is_home_page': True,
                   'is_tables_page': False,
                   'is_profile_page':False,
                   'is_all_tables_page': False,
                   'is_authenticated': is_auth,
                   'language': language_code,
                   }
        return render(request, 'home.html', context)

class ShowTable(View):
    template_name = 'Cabine.html'


    def get(self, request):
        language_code = request.path.split('/')[2]
        if request.user.is_authenticated:
            username = str(request.user.username)
            print(username)

        is_auth = request.user.username == 'admin'

        showTable  = True
        if request.user.username == 'admin':
            showTable = False
        context = {'is_home_page': False,
                   'is_tables_page': True,
                   'is_profile_page': False,
                   'is_all_tables_page': False,
                   'is_authenticated': is_auth,
                   'language':language_code}

        print(str(request.GET.get('name')))

        return render(request, self.template_name, context)


class Profile(View):
    template_name = 'Profile.html'

    def get(self, request):
        language_code = request.path.split('/')[2]
        print(language_code)
        is_auth = request.user.username == 'admin'
        context = {'is_home_page': False,
                   'is_tables_page': False,
                   'is_all_tables_page': False,
                   'is_profile_page': True,
                   'is_authenticated': is_auth,
                   'language':language_code,
                   }

        if request.user.username == 'admin':
            return render(request, 'map.html', context)
        else:
            return render(request, self.template_name, context)

def login_view(request):
    return render(request, 'registration/login.html')

class Register(View):
    template_name = 'registration/register.html'
    def get(self, request):
        context = {
            'form': UserCreationForm()
        }
        return render(request, self.template_name, context)
    def post(self, request):
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password1')
            user= authenticate(username=username, password=password)
            login(request, user)
            return redirect('home')
        context = {
            'form' : form
        }
        return render(request, self.template_name, context)

